#!/usr/bin/env python3
"""
ISO Expert Comparison Table Excel Parser

전문가가 작성한 ISO 개정안 비교표(Excel)를 구조화된 JSON으로 파싱합니다.
병합 셀이 많은 Excel 형식을 처리하며, 변경 유형을 정규화합니다.

Usage:
    python excel_parser.py --input "개정안비교표.xlsx" --sheet "8178-1" -o excel_parsed.json
    python excel_parser.py --input "개정안비교표.xlsx" --all -o excel_parsed.json

Excel 구조 (병합셀):
    Column A: 순번
    Column B: 조항 번호 (non-None = 항목 경계)
    Column C-H: IS(현행) 텍스트 (6열 병합)
    Column I-N: DIS(개정안) 텍스트 (6열 병합)
    Column O-S: 변경 유형/비고 (5열 병합)
"""

import argparse
import json
import re
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("에러: openpyxl이 필요합니다. pip install openpyxl", file=sys.stderr)
    sys.exit(1)


# 변경 유형 정규화 매핑 (한국어 → 영문 코드)
CHANGE_TYPE_MAP = {
    "내용 추가": "CONTENT_ADDED",
    "내용추가": "CONTENT_ADDED",
    "내용 변경": "CONTENT_MODIFIED",
    "내용변경": "CONTENT_MODIFIED",
    "신규 항 추가": "CLAUSE_ADDED",
    "신규항 추가": "CLAUSE_ADDED",
    "신규 항추가": "CLAUSE_ADDED",
    "신규항추가": "CLAUSE_ADDED",
    "항 순서 변경": "CLAUSE_REORDERED",
    "항 순서변경": "CLAUSE_REORDERED",
    "항순서 변경": "CLAUSE_REORDERED",
    "항순서변경": "CLAUSE_REORDERED",
    "표 변경": "TABLE_MODIFIED",
    "표변경": "TABLE_MODIFIED",
    "표 추가": "TABLE_ADDED",
    "표추가": "TABLE_ADDED",
    "삭제": "DELETED",
    "항 삭제": "CLAUSE_DELETED",
    "항삭제": "CLAUSE_DELETED",
    "용어 변경": "TERMINOLOGY_CHANGED",
    "용어변경": "TERMINOLOGY_CHANGED",
    "참조 변경": "REFERENCE_CHANGED",
    "참조변경": "REFERENCE_CHANGED",
    "편집": "EDITORIAL",
    "편집적 변경": "EDITORIAL",
    "편집적변경": "EDITORIAL",
    "그림 변경": "FIGURE_MODIFIED",
    "그림변경": "FIGURE_MODIFIED",
    "그림 추가": "FIGURE_ADDED",
    "그림추가": "FIGURE_ADDED",
    "수식 변경": "FORMULA_MODIFIED",
    "수식변경": "FORMULA_MODIFIED",
    "부속서 추가": "ANNEX_ADDED",
    "부속서추가": "ANNEX_ADDED",
    "부속서 변경": "ANNEX_MODIFIED",
    "부속서변경": "ANNEX_MODIFIED",
    "제목 변경": "TITLE_CHANGED",
    "제목변경": "TITLE_CHANGED",
    "단위 변경": "UNIT_CHANGED",
    "단위변경": "UNIT_CHANGED",
}


def normalize_change_type(raw: str) -> str:
    """한국어 변경 유형을 정규화된 영문 코드로 변환"""
    if not raw:
        return "UNKNOWN"
    cleaned = raw.strip()
    # 정확 매칭 시도
    if cleaned in CHANGE_TYPE_MAP:
        return CHANGE_TYPE_MAP[cleaned]
    # 부분 매칭 시도
    for ko, en in CHANGE_TYPE_MAP.items():
        if ko in cleaned:
            return en
    return "OTHER"


def get_merged_cell_value(ws, row, col):
    """병합 셀에서 실제 값을 가져옴. 병합 영역의 좌상단 셀 값을 반환."""
    cell = ws.cell(row=row, column=col)
    if cell.value is not None:
        return cell.value

    # 이 셀이 병합 영역에 속하는지 확인
    for merged_range in ws.merged_cells.ranges:
        if cell.coordinate in merged_range:
            # 병합 영역의 좌상단 셀 값 반환
            return ws.cell(
                row=merged_range.min_row,
                column=merged_range.min_col
            ).value
    return None


def collect_merged_text(ws, start_row, end_row, col_start, col_end):
    """지정된 행/열 범위에서 텍스트를 수집하여 합침.
    병합 셀이 여러 행에 걸쳐 있으므로, 각 행의 시작 열에서 값을 가져옴."""
    texts = []
    seen = set()

    for r in range(start_row, end_row + 1):
        for c in range(col_start, col_end + 1):
            val = get_merged_cell_value(ws, r, c)
            if val is not None:
                text = str(val).strip()
                if text and text not in seen:
                    seen.add(text)
                    texts.append(text)
    return '\n'.join(texts).strip()


def find_entry_boundaries(ws, max_row):
    """Column B에서 non-None 값을 찾아 항목 경계를 반환.
    Returns: [(start_row, end_row, clause_number), ...]"""
    boundaries = []
    header_row = None

    # 헤더 행 찾기 (조항, 현행, 개정안 등 텍스트가 있는 행)
    for r in range(1, min(10, max_row + 1)):
        val_b = ws.cell(row=r, column=2).value
        if val_b and ('조항' in str(val_b) or 'Clause' in str(val_b)):
            header_row = r
            break

    start_row = (header_row or 1) + 1  # 헤더 다음 행부터

    # Column B에서 값이 있는 행 = 새 항목 시작
    # 유효한 조항 번호 패턴: 숫자(3.5, 5.1.1), 부속서(A.1), 단일 숫자(1, 2)
    clause_num_re = re.compile(
        r'^(?:[A-Z]\.\d+(?:\.\d+)*|\d+(?:\.\d+)*)$'
    )
    entry_starts = []
    for r in range(start_row, max_row + 1):
        val = ws.cell(row=r, column=2).value
        if val is not None:
            clause_num = str(val).strip()
            if clause_num and clause_num_re.match(clause_num):
                entry_starts.append((r, clause_num))

    # 각 항목의 시작/끝 행 결정
    for i, (row, clause_num) in enumerate(entry_starts):
        if i + 1 < len(entry_starts):
            end_row = entry_starts[i + 1][0] - 1
        else:
            end_row = max_row
        boundaries.append((row, end_row, clause_num))

    return boundaries


def parse_sheet(ws, sheet_name: str) -> list:
    """단일 시트를 파싱하여 항목 리스트 반환"""
    max_row = ws.max_row
    if max_row is None or max_row < 2:
        return []

    boundaries = find_entry_boundaries(ws, max_row)
    entries = []

    for start_row, end_row, clause_num in boundaries:
        # Column A: 순번
        seq = get_merged_cell_value(ws, start_row, 1)
        seq_num = int(seq) if seq and str(seq).strip().isdigit() else None

        # Column C-H (3-8): IS(현행) 텍스트
        is_text = collect_merged_text(ws, start_row, end_row, 3, 8)

        # Column I-N (9-14): DIS(개정안) 텍스트
        dis_text = collect_merged_text(ws, start_row, end_row, 9, 14)

        # Column O-S (15-19): 변경 유형/비고
        change_raw = collect_merged_text(ws, start_row, end_row, 15, 19)

        # 변경 유형 정규화
        change_type = normalize_change_type(change_raw)

        entry = {
            "sequence": seq_num,
            "clause_number": clause_num,
            "is_text": is_text,
            "dis_text": dis_text,
            "change_description": change_raw,
            "change_type": change_type,
            "source_rows": f"{start_row}-{end_row}",
        }
        entries.append(entry)

    return entries


def parse_excel(excel_path: str, sheet_name: str = None) -> dict:
    """Excel 파일을 파싱하여 구조화된 dict 반환"""
    path = Path(excel_path)
    if not path.exists():
        raise FileNotFoundError(f"파일을 찾을 수 없습니다: {excel_path}")

    wb = openpyxl.load_workbook(str(path), data_only=True)

    result = {
        "source_file": path.name,
        "sheets": {},
    }

    if sheet_name:
        # 특정 시트만 파싱
        sheets_to_parse = []
        for name in wb.sheetnames:
            if sheet_name.lower() in name.lower():
                sheets_to_parse.append(name)
        if not sheets_to_parse:
            raise ValueError(
                f"시트 '{sheet_name}'을 찾을 수 없습니다. "
                f"사용 가능한 시트: {wb.sheetnames}"
            )
    else:
        sheets_to_parse = wb.sheetnames

    for name in sheets_to_parse:
        ws = wb[name]
        entries = parse_sheet(ws, name)
        result["sheets"][name] = {
            "sheet_name": name,
            "total_entries": len(entries),
            "entries": entries,
        }

    # 전체 통계
    total = sum(s["total_entries"] for s in result["sheets"].values())
    result["total_entries"] = total

    # 변경 유형별 통계
    type_counts = {}
    for sheet_data in result["sheets"].values():
        for entry in sheet_data["entries"]:
            ct = entry["change_type"]
            type_counts[ct] = type_counts.get(ct, 0) + 1
    result["change_type_summary"] = type_counts

    wb.close()
    return result


def main():
    parser = argparse.ArgumentParser(
        description='ISO 개정안 비교표 Excel 파싱'
    )
    parser.add_argument('--input', '-i', required=True,
                        help='Excel 파일 경로')
    parser.add_argument('--sheet', '-s',
                        help='파싱할 시트 이름 (부분 매칭). 미지정 시 전체 시트')
    parser.add_argument('--output', '-o',
                        help='출력 JSON 파일 경로 (미지정 시 stdout)')
    parser.add_argument('--pretty', action='store_true', default=True,
                        help='JSON 포맷팅 (기본값: True)')

    args = parser.parse_args()

    try:
        result = parse_excel(args.input, args.sheet)

        json_str = json.dumps(
            result, ensure_ascii=False,
            indent=2 if args.pretty else None
        )

        if args.output:
            output_path = Path(args.output)
            output_path.parent.mkdir(parents=True, exist_ok=True)
            output_path.write_text(json_str, encoding='utf-8')
            print(
                f"파싱 완료: {result['total_entries']}개 항목 "
                f"→ {args.output}"
            )
        else:
            print(json_str)

        # 요약 출력
        for name, sheet_data in result["sheets"].items():
            print(
                f"  시트 '{name}': {sheet_data['total_entries']}개 항목",
                file=sys.stderr
            )

    except Exception as e:
        print(f"에러: {e}", file=sys.stderr)
        sys.exit(1)


if __name__ == '__main__':
    main()
